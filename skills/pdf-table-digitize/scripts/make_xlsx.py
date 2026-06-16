"""
make_xlsx.py — Универсальный шаблон генерации .xlsx из данных нагрузки/таблиц.
Адаптируй данные в переменных TABLE1_DATA и TABLE2_DATA под конкретный PDF.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Стили ──────────────────────────────────────────────
thin   = Side(style="thin")
BLUE   = "BDD7EE"   # Заголовки
YELLOW = "FFF2CC"   # ИТОГО
GRAY   = "D9D9D9"   # Промежуточные итоги

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hfont(bold=True, size=10):
    return Font(name="Calibri", bold=bold, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def border_range(ws, min_r, max_r, min_c, max_c):
    b = Border(top=thin, bottom=thin, left=thin, right=thin)
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            ws.cell(r, c).border = b

def cell(ws, r, c, value, font=None, align=None, bg=None, nfmt=None):
    cl = ws.cell(row=r, column=c, value=value)
    if font:  cl.font  = font
    if align: cl.alignment = align
    if bg:    cl.fill  = fill(bg)
    if nfmt:  cl.number_format = nfmt
    return cl


# ── Данные (замени под свой PDF) ────────────────────────
# Лист 1: Аудиторная нагрузка
SHEET1_TITLE = "Аудиторная нагрузка — Фролова К.В."

# Заголовки уровень 1 (col_index, text, merge_end_col или None)
SHEET1_HEADERS_L1 = [
    (1, "ФИО сотрудника",        None),
    (2, "Дисциплина / МДК",       None),
    (3, "№ группы",               None),
    (4, "Кол-во часов",           6),   # merge D:F
    (7, "Всего",                  None),
    (8, "Экзамены",               9),   # merge H:I
    (10,"Курсовые",               None),
]
SHEET1_HEADERS_L2 = [
    (4, "1 п/г"), (5, "2 п/г"), (6, "Консульт."),
    (8, "1 сем"), (9, "2 сем"),
]
SHEET1_COL_WIDTHS = [18, 44, 14, 8, 8, 10, 8, 12, 12, 11]

# Строки данных: (ФИО, Дисциплина, Группа, 1п/г, 2п/г, Консульт, Всего, Экз1, Экз2, Курс)
# "" = объединённая ячейка (значение из первой строки)
SHEET1_DATA = [
    ("Фролова К.В.", "Информатика",                                                       "11СР(1)",  34, 36, 0,  70,  "",  "",  ""),
    ("",             "",                                                                    "12СР(1)",  34, 36, 0,  70,  "",  "",  ""),
    ("",             "",                                                                    "13ИСд(1)", 52, 88, 0, 140,  "",  18,  ""),
    ("",             "",                                                                    "21СР(1)",  55,  0,10,  65,  "",  "",  ""),
    ("",             "",                                                                    "11ТМ(1)",  39, 38, 0,  77,  "",  "",  ""),
    ("",             "",                                                                    "12ТМ(1)",  39, 38, 0,  77,  "",  "",  ""),
    ("",             "",                                                                    "11ДОУ(1)", 16, 20, 0,  36,  "",  "",  ""),
    ("",             "",                                                                    "12ДОУ(1)", 16, 20, 0,  36,  "",  "",  ""),
    ("",             "",                                                                    "11М(1)",   34, 36, 0,  70,  "",  "",  ""),
    ("",             "Проектный менеджмент в разработке и сопровождении ПО",               "13ИСд",    34, 70, 0, 104,  "",   6,  ""),
    ("",             "Прикладные компьютерные программы в профессиональной деятельности",  "21СР",      0, 44, 2,  46,  "",  "",  ""),
]
# Для столбца A объединить все строки; для B — первые N строк одной дисциплины
SHEET1_MERGE_A = True          # объединить ФИО по всем строкам
SHEET1_MERGE_B_ROWS = 9        # первые 9 строк — "Информатика"
SHEET1_ITOGO = [353, 426, 12, 791, "", 24, 0]  # значения начиная с col 4

# Лист 2: Практики
SHEET2_TITLE   = "Практики — Фролова К.А."
SHEET2_HEADERS_L1 = [
    (1, "№ п/п", None), (2, "ФИО", None),
    (3, "Наименование ПМ", None), (4, "Вид практики", None),
    (5, "Курс", None), (6, "№ группы", None),
    (7, "Кол-во часов", 9),
    (10, "Всего", None), (11, "Ставка", None), (12, "Место", None),
]
SHEET2_HEADERS_L2 = [(7, "1 п/г"), (8, "2 п/г"), (9, "Консульт.")]
SHEET2_COL_WIDTHS = [6, 16, 50, 26, 7, 18, 8, 8, 10, 8, 8, 16]

SHEET2_DATA = [
    (44, "Фролова К.А.", "ПМ.01 Разработка модулей ПО для КС",                  "УП.01 Учебная практика",          13, "23ИСд-27ИС(2)", "",   72, "", 72,  1,   "Ярославская"),
    ("",  "",             "ПМ.04 Осуществление интеграции программных модулей",  "УП.04 Учебная практика",          13, "23ИСд-27ИС(2)", 36,  36, "", 72,  1,   "Ярославская"),
    ("",  "",             'ПМ.12 Выполнение работ — Оператор ЭВМ',              "УП.12 Учебная практика",          13, "23ИСд-27ИС(2)", 72,  36, "", 108, 1,   "Ярославская"),
    ("",  "",             "ПМ.05 Проектирование и разработка ИС",               "УП.05 Учебная практика",          12, "31ИС(2)",       36,  "",  "", 36,  1,   "Шенкурский"),
    ("",  "",             "ПМ.15 Разработка и администрирование систем",        "УП.15 Учебная практика",          12, "31ИС(2)",       36,  "",  "", 36,  1,   "Шенкурский"),
    ("",  "",             "ПМ.03 Ревьюирование программных продуктов",          "УП.03 Учебная практика",          13, "32ИСд(2)",      "",  36,  "", 36,  1,   "Ярославская"),
    ("",  "",             "ПМ.05 Проектирование и разработка ИС",               "УП.05 Учебная практика",          13, "32ИСд(2)",      36,  36,  "", 72,  1,   "Ярославская"),
    ("",  "",             "ПМ.06 Сопровождение информационных систем",          "УП.06 Учебная практика",          13, "32ИСд(2)",      "",  72,  "", 72,  1,   "Ярославская"),
    ("",  "",             "ПМ.15 Разработка и администрирование систем",        "УП.15 Учебная практика",          13, "32ИСд(2)",      36,  "",  "", 36,  1,   "Ярославская"),
    ("",  "",             "ПМ.02 Осуществление интеграции программных модулей", "ПП.02 Производственная практика", 25, "31ИС",          72,  "",  "", 72,  0.5, "Шенкурский"),
    ("",  "",             'ПМ.12 Выполнение работ — Оператор ЭВМ',             "ПП.12 Производственная практика", 25, "31ИС",          72,  "",  "", 72,  0.5, "Шенкурский"),
    ("",  "",             "ПМ.15 Разработка и администрирование систем",        "ПП.15 Производственная практика", 25, "31ИС",          108, "",  "", 108, 0.5, "Шенкурский"),
]
SHEET2_ITOGO = {7: 252, 8: 0, 9: 0, 10: 792}  # col → value


# ── Построение книги ────────────────────────────────────
wb = Workbook()

def build_sheet(ws, title, headers_l1, headers_l2, col_widths, data, itogo,
                merge_a=False, merge_b_rows=0, num_cols=10,
                itogo_start_col=4, left_text_cols=(2,)):
    """Универсальная функция построения листа."""

    # Ширины столбцов
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Строка 1 — заголовок
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    cell(ws, 1, 1, title, font=hfont(bold=True, size=12),
         align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 22

    # Строки 2-3 — шапка таблицы
    for (col, text, merge_to) in headers_l1:
        if merge_to:
            ws.merge_cells(f"{get_column_letter(col)}2:{get_column_letter(merge_to)}2")
        if not merge_to:
            # Span строки 2–3 для одиночных заголовков
            ws.merge_cells(f"{get_column_letter(col)}2:{get_column_letter(col)}3")
        cell(ws, 2, col, text, font=hfont(bold=True), align=center(), bg=BLUE)

    for (col, text) in headers_l2:
        cell(ws, 3, col, text, font=hfont(bold=True), align=center(), bg=BLUE)

    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 22

    # Данные
    start = 4
    for i, row in enumerate(data):
        r = start + i
        ws.row_dimensions[r].height = 30
        for c_idx, val in enumerate(row, 1):
            al = left() if c_idx in left_text_cols else center()
            nfmt = "0.0" if isinstance(val, float) else None
            cell(ws, r, c_idx, val, font=Font(name="Calibri", size=10),
                 align=al, nfmt=nfmt)

    # Merge столбец A (№ / ФИО)
    last = start + len(data) - 1
    if merge_a:
        ws.merge_cells(f"A{start}:A{last}")
        ws.cell(start, 1).value = data[0][0]
        ws.cell(start, 1).alignment = center()
        ws.cell(start, 1).font = Font(name="Calibri", size=10)
        ws.merge_cells(f"B{start}:B{last}")
        ws.cell(start, 2).value = data[0][1]
        ws.cell(start, 2).alignment = center()
        ws.cell(start, 2).font = Font(name="Calibri", size=10)
    if merge_b_rows > 0:
        ws.merge_cells(f"B{start}:B{start + merge_b_rows - 1}")
        ws.cell(start, 2).value = data[0][1]
        ws.cell(start, 2).alignment = center()
        ws.cell(start, 2).font = Font(name="Calibri", size=10)

    # Строка ИТОГО
    itogo_row = last + 1
    ws.row_dimensions[itogo_row].height = 20
    if isinstance(itogo, list):
        ws.merge_cells(f"A{itogo_row}:C{itogo_row}")
        cell(ws, itogo_row, 1, "ИТОГО", font=hfont(bold=True), align=center(), bg=YELLOW)
        for ci, val in enumerate(itogo, itogo_start_col):
            cell(ws, itogo_row, ci, val, font=hfont(bold=True), align=center(), bg=YELLOW)
    elif isinstance(itogo, dict):
        merge_end = min(itogo.keys()) - 1
        ws.merge_cells(f"A{itogo_row}:{get_column_letter(merge_end)}{itogo_row}")
        cell(ws, itogo_row, 1, "ИТОГО", font=hfont(bold=True), align=center(), bg=YELLOW)
        for col, val in itogo.items():
            cell(ws, itogo_row, col, val, font=hfont(bold=True), align=center(), bg=YELLOW)

    border_range(ws, 2, itogo_row, 1, num_cols)
    ws.freeze_panes = "A4"


ws1 = wb.active
ws1.title = "Аудиторная нагрузка"
build_sheet(ws1, SHEET1_TITLE, SHEET1_HEADERS_L1, SHEET1_HEADERS_L2,
            SHEET1_COL_WIDTHS, SHEET1_DATA, SHEET1_ITOGO,
            merge_a=False, merge_b_rows=SHEET1_MERGE_B_ROWS,
            num_cols=10, itogo_start_col=4, left_text_cols=(2,))
# Отдельно merge ФИО по всем строкам
start = 4; last = start + len(SHEET1_DATA) - 1
ws1.merge_cells(f"A{start}:A{last}")
ws1.cell(start, 1).value = SHEET1_DATA[0][0]
ws1.cell(start, 1).alignment = center()
ws1.cell(start, 1).font = Font(name="Calibri", size=10)

ws2 = wb.create_sheet("Практики")
build_sheet(ws2, SHEET2_TITLE, SHEET2_HEADERS_L1, SHEET2_HEADERS_L2,
            SHEET2_COL_WIDTHS, SHEET2_DATA, SHEET2_ITOGO,
            merge_a=True, num_cols=12, itogo_start_col=7,
            left_text_cols=(3, 4, 12))

import sys, pathlib
out_dir = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else pathlib.Path(".")
out = out_dir / "нагрузка_2025-2026.xlsx"
wb.save(str(out))
print(f"Saved: {out}")
